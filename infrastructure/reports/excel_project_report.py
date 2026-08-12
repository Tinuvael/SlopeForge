"""OpenPyXL adapter for detached ProjectReport DTOs."""
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def write_project_report(report, path):
    wb=Workbook(); summary=wb.active; summary.title="Summary"
    production=sum(x.event_type=="production" for x in report.blasts); contour=len(report.blasts)-production
    production_volumes=[x.actual_volume_m3 for x in report.blasts if x.event_type=="production" and x.actual_volume_m3 is not None]
    volume=sum(production_volumes)
    mass=sum(x.actual_explosive_mass_kg for x in report.blasts if x.actual_explosive_mass_kg is not None)
    rows=[("Project",report.project),("Date range",f"{report.from_date.isoformat()} — {report.to_date.isoformat()}"),("",None),
          ("Blast Events total",len(report.blasts)),("Production blasts",production),("Contour blasts",contour),
          ("Actual production volume, m³",volume if production_volumes else None),
          ("Actual explosive mass, kg",mass if any(x.actual_explosive_mass_kg is not None for x in report.blasts) else None),
          ("Assessment Areas",len(report.assessments)),("Completed assessments",report.completed_assessments),
          ("Average DAI",report.average_dai),("Average FCI",report.average_fci)]
    for row in rows: summary.append(row)
    summary["A1"].font=Font(bold=True); summary.column_dimensions["A"].width=32; summary.column_dimensions["B"].width=28
    summary.append([]); summary.append(["Blast type","Count"]); summary.append(["Production",production]); summary.append(["Contour",contour])
    chart=BarChart(); chart.title="Blast events by type"; chart.add_data(Reference(summary,min_col=2,min_row=15,max_row=16)); chart.set_categories(Reference(summary,min_col=1,min_row=15,max_row=16)); summary.add_chart(chart,"D2")
    volumes=defaultdict(float)
    for row in report.blasts:
        if row.event_type=="production" and row.actual_volume_m3 is not None:volumes[row.domain]+=row.actual_volume_m3
    start=19; summary.cell(start,1,"Domain"); summary.cell(start,2,"Actual production volume, m³")
    if volumes:
        for domain,value in sorted(volumes.items()):summary.append([domain,value])
        chart2=BarChart(); chart2.title="Actual production volume by Domain"; chart2.add_data(Reference(summary,min_col=2,min_row=start,max_row=start+len(volumes)),titles_from_data=True); chart2.set_categories(Reference(summary,min_col=1,min_row=start+1,max_row=start+len(volumes))); summary.add_chart(chart2,"D18")
    else: summary.cell(start+1,1,"No actual production volume data")
    blasts=wb.create_sheet("Blast events"); bh=["Report date","Event date","Actual blast date","Domain","Type","Blast Event name","Production Block number","Horizon, m","Archived","Technical Card status","Actual block volume, m³","Actual explosive mass, kg","Actual drilling length, m"]
    blasts.append(bh)
    for x in report.blasts:blasts.append([x.report_date,x.event_date,x.actual_blast_date,x.domain,x.event_type.title(),x.name,x.block_number,x.horizon,x.archived,x.technical_card_status,x.actual_volume_m3,x.actual_explosive_mass_kg,x.actual_drilling_length_m])
    areas=wb.create_sheet("Assessment areas"); ah=["Assessment Area","Domain","Assessment date","Elevation interval","Active geometry revision","Evaluation status","DAI","FCI","Result quadrant","Linked production blocks","Linked contour blasts"]
    areas.append(ah)
    for x in report.assessments:areas.append([x.name,x.domain,x.assessment_date,x.elevation_interval,x.geometry_revision,x.evaluation_status,x.dai,x.fci,x.quadrant,", ".join(x.production_blocks),", ".join(x.contour_blasts)])
    for ws in (blasts,areas):
        ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
        for cell in ws[1]:cell.font=Font(bold=True); cell.fill=PatternFill("solid",fgColor="DCE6F1")
        for col in range(1,ws.max_column+1):ws.column_dimensions[get_column_letter(col)].width=min(32,max(12,max(len(str(ws.cell(r,col).value or "")) for r in range(1,ws.max_row+1))+2))
    for ws,cols in ((blasts,(1,2,3)),(areas,(3,))):
        for col in cols:
            for row in range(2,ws.max_row+1):ws.cell(row,col).number_format="yyyy-mm-dd"
    wb.save(path)


class OpenPyxlProjectReportWriter:
    def write(self, report, output_path) -> None:
        write_project_report(report, output_path)
