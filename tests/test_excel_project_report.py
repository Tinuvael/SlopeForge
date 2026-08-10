from datetime import date
from openpyxl import load_workbook
from reports.excel_project_report import write_project_report
from services.project_report_service import BlastReportRow,ProjectReport


def blast(kind,domain,volume):
    return BlastReportRow(date(2026,8,1),date(2026,8,1),None,domain,kind,kind,None,630,False,"completed",volume,None,None)


def test_summary_and_domain_chart_exclude_contour_volume(tmp_path):
    report=ProjectReport("Project",date(2026,8,1),date(2026,8,31),(blast("production","Production Domain",1000),blast("contour","Contour Domain",500)),())
    path=tmp_path/"report.xlsx"; write_project_report(report,path); workbook=load_workbook(path)
    summary=workbook["Summary"]
    assert summary["B7"].value==1000
    assert summary["A20"].value=="Production Domain" and summary["B20"].value==1000
    assert all(summary.cell(row,1).value!="Contour Domain" for row in range(20,summary.max_row+1))
    assert len(summary._charts)==2


def test_contour_volume_alone_does_not_become_production_volume(tmp_path):
    report=ProjectReport("Project",date(2026,8,1),date(2026,8,31),(blast("contour","Contour Domain",500),),())
    path=tmp_path/"report.xlsx"; write_project_report(report,path); summary=load_workbook(path)["Summary"]
    assert summary["B7"].value is None
    assert summary["A20"].value=="No actual production volume data"
    assert len(summary._charts)==1
